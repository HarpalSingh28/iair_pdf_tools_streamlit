import contextlib
import io
import os
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Callable, Dict, Optional, Tuple

import openpyxl
import streamlit as st

from tools.assignment.generate_pdfs import generate as generate_assignments
from tools.assessment.generate_assessments import render_all as generate_assessments
from tools.project.generate_projects import generate as generate_projects

APP_TITLE = "iAIR PDF Tools"
MAX_UPLOAD_MB = 25
BASE_DIR = Path(__file__).resolve().parent


TOOLS: Dict[str, Dict[str, object]] = {
    "Assignment PDF Generator": {
        "status": "Active",
        "description": "Upload Assignment Excel and generate one PDF per assignment row.",
        "sheet": "Assignments",
        "sample": BASE_DIR / "tools" / "assignment" / "Assignment_Input_Template.xlsx",
        "runner": generate_assignments,
        "filename": "Assignment_Input_Template.xlsx",
        "zip_name": "assignment_pdfs.zip",
    },
    "Assessment PDF Generator": {
        "status": "Active",
        "description": "Upload Assessment Excel and generate one PDF per assessment row.",
        "sheet": "Assessments",
        "sample": BASE_DIR / "tools" / "assessment" / "Assessment_Input_Template.xlsx",
        "runner": generate_assessments,
        "filename": "Assessment_Input_Template.xlsx",
        "zip_name": "assessment_pdfs.zip",
    },
    "Project PDF Generator": {
        "status": "Active",
        "description": "Upload Project Excel and generate one PDF per project row.",
        "sheet": "Projects",
        "sample": BASE_DIR / "tools" / "project" / "Project_Input_Template.xlsx",
        "runner": generate_projects,
        "filename": "Project_Input_Template.xlsx",
        "zip_name": "project_pdfs.zip",
    },
    "Question Bank Generator": {
        "status": "Coming soon",
        "description": "Placeholder for the upcoming question-bank generator.",
        "sheet": "To be added",
        "sample": None,
        "runner": None,
        "filename": "",
        "zip_name": "question_bank_pdfs.zip",
    },
}


BROWSER_ARGS = [
    "--no-sandbox",
    "--disable-setuid-sandbox",
    "--disable-dev-shm-usage",
    "--disable-gpu",
    "--single-process",
    "--no-zygote",
    "--disable-software-rasterizer",
]


def ensure_playwright_chromium() -> str:
    """Install Playwright's managed Chromium if it is not already cached.

    Streamlit Cloud may have a system Chromium, but Playwright's managed browser is
    more reliable for PDF rendering. This runs only when the cache is missing.
    """
    cache_dir = Path.home() / ".cache" / "ms-playwright"
    if cache_dir.exists() and any(cache_dir.glob("chromium-*")):
        return "Playwright Chromium already available."

    result = subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(
            "Could not install Playwright Chromium. Logs:\n" + result.stdout[-4000:]
        )
    return result.stdout[-4000:] or "Installed Playwright Chromium."


def get_secret_password() -> str:
    """Read password from Streamlit secrets or environment variable.

    Supported formats:
    1. APP_PASSWORD = "your-password"
    2. [auth]\n   password = "your-password"
    3. Environment variable APP_PASSWORD
    """
    try:
        if "APP_PASSWORD" in st.secrets:
            return str(st.secrets["APP_PASSWORD"])
    except Exception:
        pass

    try:
        return str(st.secrets["auth"]["password"])
    except Exception:
        return os.environ.get("APP_PASSWORD", "")


def require_login() -> bool:
    expected_password = get_secret_password()
    if not expected_password:
        st.warning(
            "No app password is configured. Add APP_PASSWORD in Streamlit secrets before using this publicly."
        )
        return True

    if st.session_state.get("authenticated"):
        return True

    st.title(APP_TITLE)
    st.subheader("Internal access")
    entered = st.text_input("Password", type="password", placeholder="Enter app password")
    if st.button("Login", type="primary"):
        if entered == expected_password:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False


def zip_pdf_outputs(output_dir: Path) -> Tuple[bytes, int]:
    pdf_files = sorted(output_dir.glob("*.pdf"))
    if not pdf_files:
        raise FileNotFoundError("No PDF files were generated. Please check that the Excel has valid rows.")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for pdf_file in pdf_files:
            zf.write(pdf_file, arcname=pdf_file.name)
    zip_buffer.seek(0)
    return zip_buffer.getvalue(), len(pdf_files)


def render_template_download(tool_config: Dict[str, object]) -> None:
    sample_path: Optional[Path] = tool_config.get("sample")  # type: ignore[assignment]
    if sample_path and sample_path.exists():
        st.download_button(
            label="Download Excel Template",
            data=sample_path.read_bytes(),
            file_name=str(tool_config.get("filename") or sample_path.name),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def workbook_sheetnames(uploaded_file) -> list[str]:
    data = bytes(uploaded_file.getbuffer())
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    return list(wb.sheetnames)


def validate_uploaded_workbook(uploaded_file, expected_sheet: str) -> tuple[bool, str]:
    try:
        sheets = workbook_sheetnames(uploaded_file)
    except Exception as exc:
        return False, f"Could not read Excel file. Please upload a valid .xlsx file. Details: {exc}"

    if expected_sheet not in sheets:
        return (
            False,
            f"Wrong Excel template selected. This tool expects a sheet named '{expected_sheet}', "
            f"but your file contains: {', '.join(sheets)}."
        )
    return True, ""


def generate_zip(tool_name: str, uploaded_file, runner: Callable[[str, str], None]) -> tuple[bytes, str, int]:
    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        safe_input_name = Path(uploaded_file.name).name or "input.xlsx"
        input_path = tmp_dir / safe_input_name
        output_dir = tmp_dir / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        input_path.write_bytes(uploaded_file.getbuffer())

        logs_buffer = io.StringIO()
        with contextlib.redirect_stdout(logs_buffer), contextlib.redirect_stderr(logs_buffer):
            browser_log = ensure_playwright_chromium()
            print(browser_log)
            runner(str(input_path), str(output_dir))

        zip_bytes, pdf_count = zip_pdf_outputs(output_dir)
        return zip_bytes, logs_buffer.getvalue(), pdf_count


def clear_previous_output_when_tool_changes(tool_name: str) -> None:
    if st.session_state.get("active_tool") != tool_name:
        for key in ("last_zip", "last_zip_name", "last_logs", "last_count"):
            st.session_state.pop(key, None)
        st.session_state["active_tool"] = tool_name


def render_app() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="📄", layout="centered")

    st.markdown(
        """
        <style>
            .main .block-container { padding-top: 2rem; max-width: 920px; }
            .tool-card { border: 1px solid #e5e7eb; border-radius: 14px; padding: 1rem; background: #fafafa; }
            .small-muted { color: #6b7280; font-size: 0.92rem; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if not require_login():
        return

    st.title(APP_TITLE)
    st.caption("Internal Excel to PDF automation portal")

    with st.sidebar:
        st.header("Tools")
        tool_name = st.radio("Select generator", list(TOOLS.keys()), label_visibility="collapsed")
        st.divider()
        st.caption("Active tools: Assignment, Assessment, and Project. Question Bank can be added later in the same app.")
        if st.session_state.get("authenticated") and st.button("Logout"):
            st.session_state.clear()
            st.rerun()

    clear_previous_output_when_tool_changes(tool_name)
    tool_config = TOOLS[tool_name]
    status = tool_config["status"]
    runner = tool_config["runner"]
    expected_sheet = str(tool_config["sheet"])

    st.subheader(tool_name)
    st.markdown(
        f"""
        <div class="tool-card">
            <b>Status:</b> {status}<br>
            <span class="small-muted">{tool_config['description']}</span><br><br>
            <b>Expected Excel sheet:</b> {expected_sheet}
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.write("")
    render_template_download(tool_config)

    if runner is None:
        st.info("This tool is not added yet. Its folder and runner can be connected later.")
        return

    uploaded_file = st.file_uploader(
        "Upload completed Excel file",
        type=["xlsx"],
        help=f"Maximum recommended size: {MAX_UPLOAD_MB} MB. Please use the correct template for {tool_name}.",
    )

    file_is_valid = False
    if uploaded_file is not None:
        upload_size_mb = len(uploaded_file.getbuffer()) / (1024 * 1024)
        st.caption(f"Uploaded: {uploaded_file.name} · {upload_size_mb:.2f} MB")
        if upload_size_mb > MAX_UPLOAD_MB:
            st.error(f"File is too large. Please upload an Excel file under {MAX_UPLOAD_MB} MB.")
            return

        ok, message = validate_uploaded_workbook(uploaded_file, expected_sheet)
        if ok:
            file_is_valid = True
        else:
            st.error(message)

    generate_clicked = st.button("Generate PDF ZIP", type="primary", disabled=not file_is_valid, use_container_width=True)

    if generate_clicked and uploaded_file is not None:
        with st.spinner("Generating PDFs. First run may take a minute while Chromium is prepared..."):
            try:
                zip_bytes, logs, pdf_count = generate_zip(tool_name, uploaded_file, runner)  # type: ignore[arg-type]
                st.session_state["last_zip"] = zip_bytes
                st.session_state["last_zip_name"] = str(tool_config.get("zip_name") or "generated_pdfs.zip")
                st.session_state["last_logs"] = logs
                st.session_state["last_count"] = pdf_count
                st.success(f"Done. Generated {pdf_count} PDF file(s).")
            except Exception as exc:
                st.session_state.pop("last_zip", None)
                st.error("PDF generation failed.")
                st.exception(exc)

    if st.session_state.get("last_zip"):
        st.download_button(
            label="Download Generated ZIP",
            data=st.session_state["last_zip"],
            file_name=st.session_state.get("last_zip_name", "generated_pdfs.zip"),
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )
        with st.expander("Generation logs"):
            st.code(st.session_state.get("last_logs", "No logs."))


if __name__ == "__main__":
    render_app()
