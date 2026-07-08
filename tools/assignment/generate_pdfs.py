"""
Reads Assignment_Input_Template.xlsx (sheet: 'Assignments') and generates
one formatted assignment PDF per row into the output/ folder.

Usage:
    python3 generate_pdfs.py [input_excel_path] [output_folder]

Defaults:
    input:  Assignment_Input_Template.xlsx
    output: output/
"""
import sys
import os
import shutil
import base64
import openpyxl
from jinja2 import Environment, FileSystemLoader
from playwright.sync_api import sync_playwright

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = BASE_DIR  # assignment_template.html must sit in the same folder as this script
LOGO_PATH = os.path.join(BASE_DIR, "assets", "iair_logo.png")


def load_logo_base64():
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    print(f"WARNING: logo not found at '{LOGO_PATH}'. PDFs will be generated WITHOUT the logo.")
    print("Make sure the 'assets' folder (containing iair_logo.png) sits next to generate_pdfs.py.\n")
    return None


LOGO_B64 = load_logo_base64()


def split_pipe(value):
    if value is None:
        return []
    value = str(value).strip()
    if not value:
        return []
    return [v.strip() for v in value.split("|") if v.strip()]


def split_double(value):
    """Split a pipe-separated list where each item is 'answer :: explanation'."""
    items = split_pipe(value)
    result = []
    for item in items:
        if "::" in item:
            ans, exp = item.split("::", 1)
            result.append({"answer": ans.strip(), "explanation": exp.strip()})
        else:
            result.append({"answer": item.strip(), "explanation": ""})
    return result


def row_to_context(row_dict):
    section_a_items = [row_dict.get(f"secA_q{i}") for i in range(1, 6)]
    section_a_items = [x for x in section_a_items if x and str(x).strip()]
    section_a_answers = split_pipe(row_dict.get("secA_answers"))

    section_b_pairs = []
    section_b_answers = []
    for i in range(1, 5):
        a = row_dict.get(f"secB_colA_{i}")
        b = row_dict.get(f"secB_colB_{i}")
        if a and str(a).strip():
            section_b_pairs.append({"col_a": a, "col_b": b})
            correct_letter = row_dict.get(f"secB_correct_{i}")
            if correct_letter and str(correct_letter).strip():
                letter = str(correct_letter).strip().lower()
                explanation = row_dict.get(f"secB_explanation_{i}")
                if not explanation or not str(explanation).strip():
                    # fall back to reusing the Column B text at that letter's position
                    idx = "abcdefgh".find(letter)
                    b_texts = [row_dict.get(f"secB_colB_{j}") for j in range(1, 5)]
                    explanation = b_texts[idx] if 0 <= idx < len(b_texts) and b_texts[idx] else ""
                section_b_answers.append({"num": i, "letter": letter, "explanation": explanation})

    section_c_items = [row_dict.get(f"secC_q{i}") for i in range(1, 5)]
    section_c_items = [x for x in section_c_items if x and str(x).strip()]
    section_c_answers = split_double(row_dict.get("secC_answers"))

    # ---- Section D: AI / Human / Both tri-checkbox format ----
    section_d_items = []
    for i in range(1, 4):
        stmt = row_dict.get(f"secD_q{i}")
        if stmt and str(stmt).strip():
            section_d_items.append({"statement": stmt})
    section_d_answers = split_double(row_dict.get("secD_answers"))
    section_d_options = split_pipe(row_dict.get("secD_options"))
    if len(section_d_options) != 3:
        section_d_options = ["AI", "Human", "Both"]

    section_e_items = []
    section_e_answers = []
    for i in (1, 2):
        q = row_dict.get(f"secE_q{i}")
        if q and str(q).strip():
            subparts = split_pipe(row_dict.get(f"secE_q{i}_subparts"))
            section_e_items.append({"question": q, "subparts": subparts})
            ans = row_dict.get(f"secE_q{i}_answer")
            if ans and str(ans).strip():
                section_e_answers.append(ans)

    class_range = str(row_dict.get("class_range", "")).strip()
    org_short_name = row_dict.get("org_short_name") or "iAIR"
    assignment_title = row_dict.get("assignment_title", "")

    include_solutions = str(row_dict.get("include_solutions", "")).strip().lower() in ("yes", "y", "true", "1")

    context = {
        "subject_title": row_dict.get("subject_title", ""),
        "assignment_title": assignment_title,
        "class_range": class_range,
        "max_marks": row_dict.get("max_marks", ""),
        "time_allowed": row_dict.get("time_allowed", ""),
        "week_label": row_dict.get("week_label", ""),
        "footer_left": f"{assignment_title} | Classes {class_range} | {org_short_name}",

        "section_a_marks": row_dict.get("secA_marks", ""),
        "section_a_wordbank": split_pipe(row_dict.get("secA_wordbank")),
        "section_a_items": section_a_items,
        "section_a_answers": section_a_answers,

        "section_b_marks": row_dict.get("secB_marks", ""),
        "section_b_pairs": section_b_pairs,
        "section_b_answers": section_b_answers,

        "section_c_marks": row_dict.get("secC_marks", ""),
        "section_c_options": split_pipe(row_dict.get("secC_options")),
        "section_c_items": section_c_items,
        "section_c_answers": section_c_answers,

        "section_d_marks": row_dict.get("secD_marks", ""),
        "section_d_title": row_dict.get("secD_title", "AI, Human or Both?"),
        "section_d_intro": row_dict.get("secD_intro", ""),
        "section_d_items": section_d_items,
        "section_d_options": section_d_options,
        "section_d_answers": section_d_answers,

        "section_e_marks": row_dict.get("secE_marks", ""),
        "section_e_items": section_e_items,
        "section_e_answers": section_e_answers,

        "student_reminder": row_dict.get("student_reminder", ""),

        "show_solutions": include_solutions,

        "logo_base64": LOGO_B64,
    }
    return context


def generate(input_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb["Assignments"]

    headers = [cell.value for cell in ws[1]]
    env = Environment(loader=FileSystemLoader(TEMPLATE_DIR))
    template = env.get_template("assignment_template.html")

    rows_to_render = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not row_dict.get("assignment_id"):
            continue  # skip empty rows
        rows_to_render.append(row_dict)

    count = 0
    with sync_playwright() as p:
        executable_path = (
            os.environ.get("PLAYWRIGHT_CHROMIUM_EXECUTABLE")
            or shutil.which("chromium")
            or shutil.which("chromium-browser")
            or shutil.which("google-chrome")
            or shutil.which("google-chrome-stable")
        )
        launch_kwargs = {"args": ["--no-sandbox", "--disable-dev-shm-usage"]}
        if executable_path:
            launch_kwargs["executable_path"] = executable_path
        browser = p.chromium.launch(**launch_kwargs)
        page = browser.new_page()
        for row_dict in rows_to_render:
            context = row_to_context(row_dict)
            html_out = template.render(**context)

            assignment_id = str(row_dict["assignment_id"]).strip()
            out_path = os.path.join(output_dir, f"{assignment_id}.pdf")

            page.set_content(html_out, wait_until="load")
            page.emulate_media(media="print")

            if LOGO_B64:
                header_template = f"""
                    <div style="width:100%; margin:0; padding:0 0 0 1.5cm; -webkit-print-color-adjust:exact;">
                        <img src="data:image/png;base64,{LOGO_B64}" style="height:28px; width:auto; display:block;">
                    </div>
                """
                top_margin = "2.2cm"
            else:
                header_template = "<span></span>"
                top_margin = "1.6cm"

            page.pdf(
                path=out_path,
                format="A4",
                print_background=True,
                margin={"top": top_margin, "bottom": "2.1cm", "left": "1.5cm", "right": "1.5cm"},
                display_header_footer=True,
                header_template=header_template,
                footer_template=f"""
                    <div style="width:100%; font-size:8.5pt; font-family:Arial, sans-serif;
                                padding:0 1.5cm; display:flex; justify-content:space-between;">
                        <span style="color:#1f4e79; font-weight:bold;">{context['footer_left']}</span>
                        <span style="color:#444;">Page <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                    </div>
                """,
            )
            print(f"Generated: {out_path}")
            count += 1
        browser.close()

    print(f"\nDone. {count} assignment PDF(s) created in '{output_dir}'.")


if __name__ == "__main__":
    input_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, "Assignment_Input_Template.xlsx")
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE_DIR, "output")
    generate(input_path, output_dir)
