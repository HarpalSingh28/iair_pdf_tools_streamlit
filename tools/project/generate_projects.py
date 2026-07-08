"""
Reads Project_Input_Template.xlsx (sheet: 'Projects') and generates
one formatted IAIR project-work PDF per row into the output/ folder.

Usage:
    python generate_projects.py [input_excel_path] [output_folder]

Defaults:
    input:  Project_Input_Template.xlsx
    output: output/
"""
import base64
import html
import os
import re
import sys

import openpyxl
from jinja2 import Environment, FileSystemLoader
from playwright.sync_api import sync_playwright

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = BASE_DIR
LOGO_PATH = os.path.join(BASE_DIR, "assets", "iair_logo.png")


def load_logo_base64():
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    print(f"WARNING: logo not found at '{LOGO_PATH}'. PDFs will be generated WITHOUT the logo.")
    print("Make sure assets/iair_logo.png sits next to generate_projects.py.\n")
    return None


LOGO_B64 = load_logo_base64()


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def split_pipe(value):
    value = clean(value)
    if not value:
        return []
    return [item.strip() for item in value.split("|") if item.strip()]


def split_double(value, default_title="", default_text=""):
    """Pipe-list where each item is title :: text."""
    result = []
    for item in split_pipe(value):
        if "::" in item:
            title, text = item.split("::", 1)
        else:
            title, text = item, ""
        result.append({"title": title.strip(), "text": text.strip()})
    if not result and (default_title or default_text):
        result.append({"title": default_title, "text": default_text})
    return result


def split_table(value, columns):
    """Pipe-list where each row uses :: as cell separator."""
    rows = []
    for item in split_pipe(value):
        cells = [part.strip() for part in item.split("::")]
        cells = cells[:columns] + [""] * max(0, columns - len(cells))
        rows.append(cells)
    return rows


def split_labelled_lines(value):
    """Pipe-list where each item is label :: prompt/text."""
    return split_double(value)


def as_bool(value):
    return clean(value).lower() in {"yes", "y", "true", "1"}


def line_count(value, default=3):
    txt = clean(value)
    if not txt:
        return default
    try:
        return max(1, int(float(txt)))
    except ValueError:
        return default


def html_lines(value):
    """Escape text and keep intentional line breaks. Use semicolon+space for compact natural breaks."""
    txt = clean(value)
    txt = html.escape(txt)
    txt = txt.replace("\n", "<br>")
    return txt


def first_or_empty(items, i):
    return items[i] if i < len(items) else ""


def row_to_context(row_dict):
    class_range = clean(row_dict.get("class_range")) or "6-8"
    week_label = clean(row_dict.get("week_label")) or "Week 1"
    org_short_name = clean(row_dict.get("org_short_name")) or "iAIR"
    project_work_label = clean(row_dict.get("project_work_label")) or "Project Work"
    footer_title = clean(row_dict.get("footer_title")) or f"AI {week_label} {project_work_label}"

    step_items_1_6 = split_double(row_dict.get("steps_1_6"))
    step_items_7_12 = split_double(row_dict.get("steps_7_12"))

    flow_steps = split_double(row_dict.get("flow_steps"))
    if not flow_steps:
        flow_steps = [
            {"title": "Data Received", "text": "What information enters the system?"},
            {"title": "Data Studied", "text": "What does the system compare or examine?"},
            {"title": "Pattern Identified", "text": "What repeated trend, feature or relationship is found?"},
            {"title": "AI Result", "text": "What recognition, prediction, recommendation or response is produced?"},
            {"title": "Human Check", "text": "What should a person verify?"},
            {"title": "Human Decision or Action", "text": "What final decision or action should a person take?"},
        ]

    context = {
        "logo_base64": LOGO_B64,
        "subject_title": clean(row_dict.get("subject_title")) or "Artificial Intelligence",
        "week_label": week_label,
        "project_work_label": project_work_label,
        "class_range": class_range,
        "workbook_label": clean(row_dict.get("workbook_label")) or "Project Workbook",
        "max_marks": clean(row_dict.get("max_marks")) or "25",
        "suggested_duration": clean(row_dict.get("suggested_duration")) or "4-5 Days",
        "project_mode": clean(row_dict.get("project_mode")) or "Individual Project",
        "org_short_name": org_short_name,
        "footer_left": f"{footer_title} | Class {class_range} | {org_short_name}",

        "project_theme_label": clean(row_dict.get("project_theme_label")) or "Young AI Designer Challenge",
        "project_title": clean(row_dict.get("project_title")) or "Design a Responsible AI School Helper",
        "driving_question": clean(row_dict.get("driving_question")) or "How can we design an AI-based helper that solves a school problem while keeping humans in control of important decisions?",
        "mission_title": clean(row_dict.get("mission_title")) or "Project Mission",
        "mission_text": clean(row_dict.get("mission_text")),
        "mission_points": split_pipe(row_dict.get("mission_points")),
        "digital_safety_title": clean(row_dict.get("digital_safety_title")) or "Digital Safety Rule",
        "digital_safety_text": clean(row_dict.get("digital_safety_text")),

        "steps_1_6": step_items_1_6,
        "steps_7_12": step_items_7_12,
        "remember_note": clean(row_dict.get("remember_note")),
        "submit_items": split_pipe(row_dict.get("submit_items")),
        "work_plan": split_table(row_dict.get("work_plan"), 2),

        "learning_objectives": split_double(row_dict.get("learning_objectives")),
        "part_a_audience_options": split_pipe(row_dict.get("part_a_audience_options")),

        "human_steps": split_double(row_dict.get("human_steps")),
        "intelligence_options": split_pipe(row_dict.get("intelligence_options")),

        "automation_options_1": split_pipe(row_dict.get("automation_options_1")),
        "automation_options_2": split_pipe(row_dict.get("automation_options_2")),
        "solution_choice_options": split_pipe(row_dict.get("solution_choice_options")),
        "difference_left_heading": clean(row_dict.get("difference_left_heading")) or "Simple Automation Would...",
        "difference_right_heading": clean(row_dict.get("difference_right_heading")) or "Artificial Intelligence Would...",
        "helpful_hint": clean(row_dict.get("helpful_hint")),

        "data_type_options": split_pipe(row_dict.get("data_type_options")),
        "data_source_options": split_pipe(row_dict.get("data_source_options")),
        "private_info_lines": line_count(row_dict.get("private_info_lines"), 3),

        "result_type_options": split_pipe(row_dict.get("result_type_options")),
        "logic_hint": clean(row_dict.get("logic_hint")),

        "ai_domain_cards": split_double(row_dict.get("ai_domain_cards")),
        "flow_steps": flow_steps,
        "flow_summary": clean(row_dict.get("flow_summary")) or "DATA -> STUDY -> PATTERN -> AI RESULT -> HUMAN CHECK -> HUMAN DECISION",

        "prototype_screen1_boxes": split_pipe(row_dict.get("prototype_screen1_boxes")),
        "prototype_screen2_boxes": split_pipe(row_dict.get("prototype_screen2_boxes")),
        "prototype_screen3_boxes": split_pipe(row_dict.get("prototype_screen3_boxes")),
        "prototype_tip": clean(row_dict.get("prototype_tip")),

        "test_case1_fields": split_pipe(row_dict.get("test_case1_fields")),
        "test_case2_fields": split_pipe(row_dict.get("test_case2_fields")),
        "system_work_options": split_pipe(row_dict.get("system_work_options")),
        "mistake_cause_options": split_pipe(row_dict.get("mistake_cause_options")),

        "responsible_rules": split_pipe(row_dict.get("responsible_rules")),
        "can_user_options": split_pipe(row_dict.get("can_user_options")),
        "presentation_checklist": split_pipe(row_dict.get("presentation_checklist")),
        "reflection_q1_3": split_pipe(row_dict.get("reflection_q1_3")),
        "reflection_q4_7": split_pipe(row_dict.get("reflection_q4_7")),
        "submission_checklist": split_pipe(row_dict.get("submission_checklist")),
        "rubric_rows": split_table(row_dict.get("rubric_rows"), 2),
        "teacher_guidance_note": clean(row_dict.get("teacher_guidance_note")),
    }
    return context


def generate(input_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if "Projects" not in wb.sheetnames:
        raise ValueError("Input workbook must contain a sheet named 'Projects'.")
    ws = wb["Projects"]

    headers = [cell.value for cell in ws[1]]
    env = Environment(loader=FileSystemLoader(TEMPLATE_DIR), autoescape=True)
    env.filters["html_lines"] = html_lines
    template = env.get_template("project_template.html")

    rows_to_render = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not clean(row_dict.get("project_id")):
            continue
        rows_to_render.append(row_dict)

    count = 0
    with sync_playwright() as p:
        
        launch_kwargs = {
            "headless": True,
            "args": [
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--single-process",
                "--no-zygote",
                "--disable-software-rasterizer",
            ],
        }
        browser = p.chromium.launch(**launch_kwargs)
        page = browser.new_page(viewport={"width": 1240, "height": 1754})
        for row_dict in rows_to_render:
            context = row_to_context(row_dict)
            html_out = template.render(**context)

            project_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", clean(row_dict["project_id"]))
            out_path = os.path.join(output_dir, f"{project_id}.pdf")

            page.set_content(html_out, wait_until="load")
            page.emulate_media(media="print")
            page.pdf(
                path=out_path,
                format="A4",
                print_background=True,
                margin={"top": "1.15cm", "bottom": "1.75cm", "left": "1.35cm", "right": "1.35cm"},
                display_header_footer=True,
                header_template="<span></span>",
                footer_template=f"""
                    <div style="width:100%; font-size:8.2pt; font-family:Arial, sans-serif;
                                padding:0 1.35cm; display:flex; justify-content:space-between;">
                        <span style="color:#1f5d8c; font-weight:bold;">{html.escape(context['footer_left'])}</span>
                        <span style="color:#56606b;">Page <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                    </div>
                """,
            )
            print(f"Generated: {out_path}")
            count += 1
        browser.close()

    print(f"\nDone. {count} project PDF(s) created in '{output_dir}'.")


if __name__ == "__main__":
    input_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, "Project_Input_Template.xlsx")
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE_DIR, "output")
    generate(input_path, output_dir)
