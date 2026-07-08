"""
Reads Assessment_Input_Template.xlsx (sheet: 'Assessments') and generates
one formatted assessment PDF per row into the output/ folder.

Usage:
    python generate_assessments.py [input_excel_path] [output_folder]

Defaults:
    input:  Assessment_Input_Template.xlsx
    output: output/

Stack: openpyxl + Jinja2 + Playwright Chromium.
Do not use WeasyPrint.
"""
import base64
import os
import re
import sys
import shutil
from typing import Any, Dict, List

import openpyxl
from jinja2 import Environment, FileSystemLoader, select_autoescape
from playwright.sync_api import sync_playwright

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = BASE_DIR
TEMPLATE_NAME = "assessment_template.html"
LOGO_PATH = os.path.join(BASE_DIR, "assets", "iair_logo.png")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def yes(value: Any) -> bool:
    return clean(value).lower() in {"yes", "y", "true", "1", "include", "included"}


def split_pipe(value: Any) -> List[str]:
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def to_int(value: Any, default: int) -> int:
    try:
        return int(float(clean(value)))
    except Exception:
        return default


def load_logo_base64() -> str:
    if os.path.exists(LOGO_PATH):
        with open(LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    print(f"WARNING: logo not found at '{LOGO_PATH}'. PDFs will be generated WITHOUT the logo.")
    print("Make sure assets/iair_logo.png sits next to generate_assessments.py inside the assets folder.\n")
    return ""


def option_text(row: Dict[str, Any], i: int, letter: str) -> str:
    return clean(row.get(f"mcq{i}_option_{letter.lower()}"))


def build_mcqs(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    mcqs = []
    for i in range(1, 11):
        stem = clean(row.get(f"mcq{i}_stem"))
        question = clean(row.get(f"mcq{i}_question"))
        if not stem and not question:
            continue
        opts = []
        for letter in "ABCD":
            text = option_text(row, i, letter)
            if text:
                opts.append({"letter": letter, "text": text})
        ans_letter = clean(row.get(f"mcq{i}_answer_letter")).upper()
        ans_text = option_text(row, i, ans_letter) if ans_letter in "ABCD" else clean(row.get(f"mcq{i}_answer_text"))
        if not ans_text:
            ans_text = clean(row.get(f"mcq{i}_answer_text"))
        mcqs.append({
            "number": i,
            "stem": stem,
            "question": question,
            "options": opts,
            "answer_letter": ans_letter,
            "answer_text": ans_text,
            "answer_explanation": clean(row.get(f"mcq{i}_answer_explanation")),
        })
    return mcqs



def format_vsq_solution(answer: str) -> Dict[str, Any]:
    """Return a small structured representation so solution cards can match the reference layout."""
    text = clean(answer)
    result: Dict[str, Any] = {
        "solution_intro": "",
        "solution_points": [],
        "solution_items": [],
        "solution_text": "",
    }
    if not text:
        return result

    # Q11-style: Intro: item; item; item
    if text.lower().startswith("accept") and ":" in text and ";" in text:
        intro, rest = text.split(":", 1)
        result["solution_intro"] = intro.strip() + ":"
        points = []
        for raw in rest.split(";"):
            raw = raw.strip()
            if not raw:
                continue
            if raw[-1] not in ".!?":
                raw += "."
            m = re.match(r"^(She\s+\w+)(.*)$", raw, flags=re.I)
            if m:
                points.append({"lead": m.group(1), "rest": m.group(2).strip(), "text": raw})
            else:
                points.append({"lead": "", "rest": "", "text": raw})
        result["solution_points"] = points
        return result

    # Pairs with explicit labels: Prediction/Recommendation, Capability/Limitation, etc.
    explicit_markers = ["Prediction:", "Recommendation:", "Capability:", "Limitation:"]
    present = [(m, text.find(m)) for m in explicit_markers if text.find(m) >= 0]
    if len(present) >= 2:
        present.sort(key=lambda x: x[1])
        items = []
        for idx, (marker, start) in enumerate(present):
            end = present[idx + 1][1] if idx + 1 < len(present) else len(text)
            body = text[start + len(marker):end].strip()
            items.append({"label": marker.rstrip(":"), "body": body})
        result["solution_items"] = items
        return result

    # a./b./c. style answers in a single cell.
    matches = list(re.finditer(r"(?:^|\s)([a-z]\.)\s+", text))
    if len(matches) >= 2:
        items = []
        for idx, match in enumerate(matches):
            start = match.end()
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
            segment = text[start:end].strip()
            label = match.group(1)
            body = segment
            if " - " in segment:
                head, body = segment.split(" - ", 1)
                label = f"{label} {head.strip()}"
                body = body.strip()
            else:
                label = f"{label}"
            items.append({"label": label, "body": body})
        result["solution_items"] = items
        return result

    result["solution_text"] = text
    return result

def build_vsq(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    items = []
    for idx in range(1, 6):
        question = clean(row.get(f"vsq{idx}_question"))
        if not question:
            continue
        number = clean(row.get(f"vsq{idx}_number")) or str(idx + 10)
        labels = split_pipe(row.get(f"vsq{idx}_labels"))
        solution = format_vsq_solution(clean(row.get(f"vsq{idx}_answer")))
        item = {
            "number": number,
            "question": question,
            "labels": labels,
            "line_count": to_int(row.get(f"vsq{idx}_line_count"), 3),
            "answer": clean(row.get(f"vsq{idx}_answer")),
            "marking": clean(row.get(f"vsq{idx}_marking")),
            "solution_title": clean(row.get(f"vsq{idx}_solution_title")) or question,
        }
        item.update(solution)
        items.append(item)
    return items


def case_solution_label(question: str) -> str:
    text = re.sub(r"^(which|what|why|give|mention|should|write|name)\b", "", question.strip(), flags=re.I)
    text = re.sub(r"\?$", "", text).strip()
    if not text:
        return "Answer"
    label = text.split(".")[0].strip()
    words = label.split()
    return " ".join(words[:4]).capitalize()


def build_cases(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    cases = []
    for idx in (1, 2):
        title = clean(row.get(f"case{idx}_title"))
        if not title:
            continue
        questions = []
        for letter in "abcde":
            q = clean(row.get(f"case{idx}_q_{letter}"))
            if not q:
                continue
            questions.append({
                "letter": letter,
                "question": q,
                "mark": clean(row.get(f"case{idx}_mark_{letter}")) or "1 Mark",
                "answer": clean(row.get(f"case{idx}_ans_{letter}")),
                "solution_label": clean(row.get(f"case{idx}_solution_label_{letter}")) or case_solution_label(q),
            })
        cases.append({
            "number": idx,
            "right_label": clean(row.get(f"case{idx}_right_label")) or f"Section C<br>Case Study {idx}",
            "title": title,
            "intro": clean(row.get(f"case{idx}_intro")),
            "bullets": split_pipe(row.get(f"case{idx}_bullets")),
            "pattern_sentence": clean(row.get(f"case{idx}_pattern_sentence")),
            "callouts": split_pipe(row.get(f"case{idx}_callouts")),
            "closing": clean(row.get(f"case{idx}_closing")),
            "questions": questions,
        })
    return cases


def build_marking_summary(row: Dict[str, Any]) -> List[Dict[str, str]]:
    raw = split_pipe(row.get("marking_summary_rows"))
    rows = []
    for item in raw:
        if "::" in item:
            section, marks = item.split("::", 1)
        elif ":" in item:
            section, marks = item.rsplit(":", 1)
        else:
            continue
        rows.append({"section": section.strip(), "marks": marks.strip()})
    if not rows:
        rows = [
            {"section": "Section A: Multiple-Choice Questions", "marks": "10"},
            {"section": "Section B: Very Short Answer Questions", "marks": "10"},
            {"section": "Section C: Application-Based Questions", "marks": "10"},
            {"section": "Total", "marks": clean(row.get("total_marks_label")) or "30 Marks"},
        ]
    return rows


def row_to_context(row: Dict[str, Any], logo_base64: str) -> Dict[str, Any]:
    subject_title = clean(row.get("subject_title")) or "ARTIFICIAL INTELLIGENCE"
    assessment_title = clean(row.get("assessment_title")) or "COMPREHENSIVE ASSESSMENT"
    week_label = clean(row.get("week_label")) or "WEEK 1"
    class_range = clean(row.get("class_range")) or "6-8"
    org_short_name = clean(row.get("org_short_name")) or "iAIR"
    footer_left = clean(row.get("footer_left")) or f"AI {week_label.title()} Comprehensive Assessment | Classes {class_range} | {org_short_name}"

    context = {
        "logo_base64": logo_base64,
        "subject_title": subject_title.upper(),
        "assessment_title": assessment_title.upper(),
        "week_label": week_label.upper(),
        "class_range": class_range,
        "max_marks": clean(row.get("max_marks")) or "30",
        "time_allowed": clean(row.get("time_allowed")) or "40 Minutes",
        "footer_left": footer_left,
        "general_instructions": split_pipe(row.get("general_instructions")),
        "section_a_marks": clean(row.get("section_a_marks")) or "10 x 1 = 10 Marks",
        "section_a_continued_badge": clean(row.get("section_a_continued_badge")) or "Questions 6-10",
        "section_a_remember": clean(row.get("section_a_remember")),
        "secA_solution_marks": clean(row.get("secA_solution_marks")) or "10 Marks",
        "mcqs": build_mcqs(row),
        "secB_page_title": clean(row.get("secB_page_title")) or "VERY SHORT ANSWER QUESTIONS",
        "secB_right_label": clean(row.get("secB_right_label")) or "Section B<br>10 Marks",
        "secB_marks_label": clean(row.get("secB_marks_label")) or "5 x 2 = 10 Marks",
        "secB_solution_page_title": clean(row.get("secB_solution_page_title")) or "SECTION B - VERY SHORT ANSWERS",
        "secB_solution_marks": clean(row.get("secB_solution_marks")) or "10 Marks",
        "secB_guidance": clean(row.get("secB_guidance")),
        "very_short_questions": build_vsq(row),
        "secC_page_title": clean(row.get("secC_page_title")) or "APPLICATION-BASED QUESTIONS",
        "secC_marks_label": clean(row.get("secC_marks_label")) or "2 x 5 = 10 Marks",
        "secC_solution_page_title": clean(row.get("secC_solution_page_title")) or "SECTION C - CASE STUDY SOLUTIONS",
        "secC_solution_marks": clean(row.get("secC_solution_marks")) or "10 Marks",
        "cases": build_cases(row),
        "checklist_title": clean(row.get("checklist_title")) or "Check Before Submitting",
        "checklist_items": split_pipe(row.get("checklist_items")),
        "end_label": clean(row.get("end_label")) or "END OF STUDENT ASSESSMENT",
        "show_solutions": yes(row.get("include_solutions")),
        "marking_summary": build_marking_summary(row),
        "total_marks_label": clean(row.get("total_marks_label")) or "Total: 30 Marks",
        "teacher_note": split_pipe(row.get("teacher_note")),
    }
    return context


def get_rows(input_path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if "Assessments" not in wb.sheetnames:
        raise ValueError("Workbook must contain a sheet named 'Assessments'.")
    ws = wb["Assessments"]
    headers = [clean(cell.value) for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if clean(data.get("assessment_id")):
            rows.append(data)
    return rows


def render_all(input_path: str, output_dir: str) -> None:
    os.makedirs(output_dir, exist_ok=True)
    logo_base64 = load_logo_base64()
    rows = get_rows(input_path)

    env = Environment(
        loader=FileSystemLoader(TEMPLATE_DIR),
        autoescape=select_autoescape(["html", "xml"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )
    env.globals["range"] = range
    template = env.get_template(TEMPLATE_NAME)

    with sync_playwright() as p:
        launch_kwargs = {
            "headless": True,
            "args": [
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--single-process",
                "--no-zygote",
                "--disable-software-rasterizer",
            ],
        }

        browser = p.chromium.launch(**launch_kwargs)
        page = browser.new_page(viewport={"width": 794, "height": 1123})
        for row in rows:
            context = row_to_context(row, logo_base64)
            html = template.render(**context)
            out_name = clean(row["assessment_id"]) + ".pdf"
            out_path = os.path.join(output_dir, out_name)
            page.set_content(html, wait_until="load")
            page.emulate_media(media="print")
            page.pdf(
                path=out_path,
                format="A4",
                print_background=True,
                display_header_footer=True,
                header_template="<span></span>",
                footer_template=f"""
                    <div style="width:100%; font-size:8.2pt; font-family:Arial, sans-serif;
                                padding:0 1.45cm; display:flex; justify-content:space-between; align-items:center;">
                        <span style="color:#1f5d93; font-weight:700;">{context['footer_left']}</span>
                        <span style="color:#4b5563;">Page <span class="pageNumber"></span> of <span class="totalPages"></span></span>
                    </div>
                """,
                margin={"top": "0.95cm", "bottom": "1.65cm", "left": "1.05cm", "right": "1.05cm"},
                prefer_css_page_size=True,
            )
            print(f"Generated: {out_path}")
        browser.close()
    print(f"\nDone. {len(rows)} assessment PDF(s) created in '{output_dir}'.")


if __name__ == "__main__":
    input_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, "Assessment_Input_Template.xlsx")
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(BASE_DIR, "output")
    render_all(input_path, output_dir)
